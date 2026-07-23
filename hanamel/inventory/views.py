import json
from datetime import datetime, timedelta
from itertools import groupby
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, F, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from .models import BusinessProfile, Payment, Product, Sale, SaleItem
from .templatetags.money import CURRENCY


def _next_sale_number():
    """SAL-2026-0001, restarting each year."""
    year = timezone.now().year
    prefix = f"SAL-{year}-"
    last = (Sale.objects.filter(number__startswith=prefix)
            .order_by("-number").values_list("number", flat=True).first())
    seq = int(last.split("-")[-1]) + 1 if last else 1
    return f"{prefix}{seq:04d}"


@login_required
def dashboard(request):
    today = timezone.now().date()
    week_ago = timezone.now() - timedelta(days=7)

    todays = Sale.objects.filter(date__date=today)
    week = Sale.objects.filter(date__gte=week_ago)

    low_stock = (Product.objects.filter(is_active=True,
                                        stock_qty__lte=F("low_stock_threshold"))
                 .order_by("stock_qty"))

    return render(request, "inventory/dashboard.html", {
        "todays_count": todays.count(),
        "todays_total": todays.aggregate(t=Sum("total"))["t"] or Decimal("0"),
        "week_total": week.aggregate(t=Sum("total"))["t"] or Decimal("0"),
        "low_stock": low_stock,
        "recent_sales": Sale.objects.order_by("-date")[:8],
        "page": "dash",
    })


@login_required
def product_list(request):
    q = request.GET.get("q", "").strip()
    products = Product.objects.filter(is_active=True)
    if q:
        products = products.filter(Q(name__icontains=q) | Q(species__icontains=q))
    finished = products.filter(type=Product.FINISHED).order_by("name")
    return render(request, "inventory/product_list.html", {
        "timber_groups": _timber_groups(products),
        "finished": finished, "q": q, "page": "stock"})


@login_required
def sale_list(request):
    sales = Sale.objects.order_by("-date").prefetch_related("items__product")
    return render(request, "inventory/sale_list.html", {"sales": sales, "page": "sales"})


@login_required
def sale_detail(request, pk):
    sale = get_object_or_404(Sale.objects.prefetch_related("items__product"), pk=pk)
    return render(request, "inventory/sale_detail.html", {"sale": sale, "page": "sales"})


@login_required
def sale_invoice(request, pk):
    """Print-friendly invoice. Rendered from the sale record on demand —
    nothing is stored on disk, so it can never go stale."""
    sale = get_object_or_404(
        Sale.objects.prefetch_related("items__product", "payments"), pk=pk)
    return render(request, "inventory/invoice.html",
                  {"sale": sale, "biz": BusinessProfile.get()})


@login_required
def sale_create(request):
    """Record a sale. Prices are ALWAYS recomputed server-side from the DB —
    the browser's running total is a convenience, never the source of truth."""
    products = Product.objects.filter(is_active=True).order_by("type", "species", "-width")

    if request.method == "POST":
        product_ids = request.POST.getlist("product_id")
        quantities = request.POST.getlist("qty")
        customer = request.POST.get("customer_name", "").strip()
        pay_method = request.POST.get("payment_method", Payment.CASH)
        raw_paid = request.POST.get("amount_paid", "").strip()
        mpesa_ref = request.POST.get("mpesa_receipt", "").strip()

        lines = []
        errors = []
        for pid, raw_qty in zip(product_ids, quantities):
            if not pid or not raw_qty:
                continue
            try:
                qty = Decimal(str(raw_qty))
            except (InvalidOperation, ValueError):
                errors.append("One of the quantities isn't a number.")
                continue
            if qty <= 0:
                continue
            product = Product.objects.filter(pk=pid, is_active=True).first()
            if not product:
                errors.append("A product on this sale is no longer available.")
                continue
            lines.append((product, qty))

        if not lines and not errors:
            errors.append("Add at least one item before recording the sale.")

        if errors:
            for e in errors:
                messages.error(request, e)
            return render(request, "inventory/sale_form.html", {
                "timber_groups": _timber_groups(products),
                "finished": [p for p in products if p.type == Product.FINISHED],
                "catalog_json": _catalog_json(products), "methods": Payment.METHOD_CHOICES,
                "page": "new", "customer_name": customer})

        try:
            with transaction.atomic():
                # lock the rows we're about to decrement
                ids = [p.pk for p, _ in lines]
                locked = {p.pk: p for p in
                          Product.objects.select_for_update().filter(pk__in=ids)}

                short = [f"{locked[p.pk]} (have {locked[p.pk].stock_qty}, need {q})"
                         for p, q in lines if locked[p.pk].stock_qty < q]
                if short:
                    raise ValueError("Not enough stock: " + "; ".join(short))

                sale = Sale.objects.create(number=_next_sale_number(),
                                           customer_name=customer)
                for p, qty in lines:
                    prod = locked[p.pk]
                    is_timber = prod.type == Product.TIMBER
                    SaleItem.objects.create(
                        sale=sale,
                        product=prod,
                        qty=qty,
                        volume_per_piece=prod.volume_m3 if is_timber else None,
                        rate=prod.rate_per_m3 if is_timber else prod.price,
                    )
                    prod.stock_qty = prod.stock_qty - qty
                    prod.save(update_fields=["stock_qty"])
                sale.recalculate_total()

                # Record what was actually handed over. Blank means "pay in full";
                # a smaller figure leaves a balance owing (contractor deposits).
                if raw_paid == "":
                    paid = sale.total
                else:
                    try:
                        paid = Decimal(raw_paid)
                    except (InvalidOperation, ValueError):
                        paid = Decimal("0")
                if paid > 0:
                    Payment.objects.create(
                        sale=sale, method=pay_method, amount=min(paid, sale.total),
                        mpesa_receipt=mpesa_ref if pay_method == Payment.MPESA else "")
        except ValueError as exc:
            messages.error(request, str(exc))
            return render(request, "inventory/sale_form.html", {
                "timber_groups": _timber_groups(products),
                "finished": [p for p in products if p.type == Product.FINISHED],
                "catalog_json": _catalog_json(products), "methods": Payment.METHOD_CHOICES,
                "page": "new", "customer_name": customer})

        messages.success(request, f"Recorded {sale.number} — {CURRENCY} {sale.total:,.2f}")
        return redirect("sale_detail", pk=sale.pk)

    return render(request, "inventory/sale_form.html", {
        "timber_groups": _timber_groups(products),
        "finished": [p for p in products if p.type == Product.FINISHED],
        "catalog_json": _catalog_json(products), "page": "new",
        "methods": Payment.METHOD_CHOICES})


def _timber_groups(products):
    """Group timber by species so the picker stays readable with several species."""
    timber = [p for p in products if p.type == Product.TIMBER]
    timber.sort(key=lambda p: (p.species or p.name,
                               -float(p.width or 0), -float(p.thickness or 0),
                               -float(p.length or 0)))
    return [{"species": sp, "items": list(items)}
            for sp, items in groupby(timber, key=lambda p: p.species or p.name)]


def _catalog_json(products):
    """Server-computed unit prices, handed to the browser for instant totals."""
    return json.dumps([{
        "id": p.pk,
        "label": str(p),
        "type": p.type,
        "unit_price": float(p.unit_price() or 0),
        "stock": float(p.stock_qty),
        "volume": float(p.volume_m3) if p.volume_m3 else None,
    } for p in products])


# --- monthly books export -------------------------------------------------

def _month_range(year, month):
    """Timezone-aware [start, end) for a calendar month in the yard's local time."""
    tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime(year, month, 1), tz)
    end = timezone.make_aware(
        datetime(year + (month == 12), (month % 12) + 1, 1), tz)
    return start, end


@login_required
def invoice_pack(request, year, month):
    """Every invoice for a month on one page, page-broken for printing.
    Browser print gives a single PDF of the lot."""
    start, end = _month_range(year, month)
    sales = (Sale.objects.filter(date__gte=start, date__lt=end)
             .prefetch_related("items__product", "payments").order_by("date"))
    return render(request, "inventory/invoice_pack.html", {
        "sales": sales, "biz": BusinessProfile.get(),
        "label": start.strftime("%B %Y"), "page": "reports"})


@login_required
def reports(request):
    """Pick a month to export. Lists months that actually have sales."""
    months = (Sale.objects.annotate(m=TruncMonth("date"))
              .values("m").annotate(n=Count("id"), total=Sum("total"))
              .order_by("-m"))
    return render(request, "inventory/reports.html", {"months": months, "page": "reports"})


@login_required
def export_month(request, year, month):
    """One .xlsx per month: a summary sheet the accountant can post from,
    plus a line-item sheet for anyone who needs the detail."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    start, end = _month_range(year, month)
    sales = (Sale.objects.filter(date__gte=start, date__lt=end)
             .prefetch_related("items__product", "payments").order_by("date"))
    biz = BusinessProfile.get()

    wb = Workbook()
    head_font = Font(name="Arial", bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="0F9455")
    body = Font(name="Arial")
    title = Font(name="Arial", bold=True, size=13)
    money = '#,##0.00'
    thin = Side(style="thin", color="D9D9D9")
    edge = Border(bottom=thin)

    def write_header(ws, headers, row=4):
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font, cell.fill = head_font, head_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = edge

    # ---------- Sheet 1: one row per sale ----------
    ws = wb.active
    ws.title = "Sales"
    ws["A1"] = biz.name
    ws["A1"].font = title
    ws["A2"] = f"Sales for {start.strftime('%B %Y')}"
    ws["A2"].font = body
    if biz.kra_pin:
        ws["A3"] = f"KRA PIN: {biz.kra_pin}"
        ws["A3"].font = body

    cols = ["Date", "Invoice no.", "Customer", "Items",
            f"Total ({CURRENCY})", f"Paid ({CURRENCY})", f"Balance ({CURRENCY})",
            "Payment method", "M-Pesa ref"]
    write_header(ws, cols)

    r = 5
    for s in sales:
        refs = ", ".join(p.mpesa_receipt for p in s.payments.all() if p.mpesa_receipt)
        values = [s.date.strftime("%Y-%m-%d %H:%M"), s.number,
                  s.customer_name or "Walk-in", s.items.count(),
                  float(s.total), float(s.amount_paid), float(s.balance),
                  s.payment_summary, refs]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = body
            if c in (5, 6, 7):
                cell.number_format = money
        r += 1

    if sales:
        for c, label in ((4, "TOTAL"), (5, None), (6, None), (7, None)):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Arial", bold=True)
            if label:
                cell.value = label
                cell.alignment = Alignment(horizontal="right")
            else:
                letter = cell.column_letter
                cell.value = f"=SUM({letter}5:{letter}{r-1})"
                cell.number_format = money

    for col, w in zip("ABCDEFGHI", (17, 15, 24, 7, 14, 14, 14, 16, 16)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

    # ---------- Sheet 2: one row per line item ----------
    ws2 = wb.create_sheet("Line items")
    ws2["A1"] = f"Line detail — {start.strftime('%B %Y')}"
    ws2["A1"].font = title
    cols2 = ["Date", "Invoice no.", "Customer", "Item", "Species", "Size",
             "Volume m³/pc", "Qty", f"Rate ({CURRENCY})", f"Amount ({CURRENCY})"]
    write_header(ws2, cols2)

    r = 5
    for s in sales:
        for it in s.items.all():
            p = it.product
            size = (f"{p.width:g}x{p.thickness:g}x{p.length:g}"
                    if p.type == Product.TIMBER and p.width else "")
            values = [s.date.strftime("%Y-%m-%d"), s.number,
                      s.customer_name or "Walk-in", p.name, p.species or "", size,
                      float(it.volume_per_piece) if it.volume_per_piece else "",
                      float(it.qty), float(it.rate), float(it.amount)]
            for c, v in enumerate(values, start=1):
                cell = ws2.cell(row=r, column=c, value=v)
                cell.font = body
                if c in (9, 10):
                    cell.number_format = money
                if c == 7 and v != "":
                    cell.number_format = '0.00000'
            r += 1
    if r > 5:
        cell = ws2.cell(row=r, column=9, value="TOTAL")
        cell.font = Font(name="Arial", bold=True)
        cell.alignment = Alignment(horizontal="right")
        tot = ws2.cell(row=r, column=10, value=f"=SUM(J5:J{r-1})")
        tot.font = Font(name="Arial", bold=True)
        tot.number_format = money

    for col, w in zip("ABCDEFGHIJ", (12, 15, 22, 16, 13, 13, 13, 8, 13, 14)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A5"

    filename = f"{biz.name.replace(' ', '_')}_sales_{year}_{month:02d}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response